#импорт библиотек
import glob
from pathlib import Path, PureWindowsPath
import random
import openpyxl
from openpyxl.utils import get_column_letter
import re
import os, os.path
from pyzbar.pyzbar import decode, ZBarSymbol
from PIL import Image
from playsound import playsound
import cv2

#переменные______________________________________________________________________
correct_path = r"C:\\Users\\EreminAV01\\PycharmProjects\\pythonProject6\\vihod\\"
path = Path(correct_path)
ooo = 0
per = random.randint(0, 10000)
cheredovanie = 1

endpic = 'pic/endpic.png'

o1 = str(1)
o2 = str(2)
o3 = str(3)
o4 = str(4)
o5 = str(5)
o6 = str(6)
o7 = str(7)
o8 = str(8)
o9 = str(9)

#тело кода__________________________________________________________________________________
def o():

#блок обхода папки__________________________________________________________________
    for filename in glob.glob("vhod00001" + "/*." + "jpeg" or "jpg"): # обходим заданную папку
            if str(filename): # если имя
                print('Печатаем распознанный путь:',filename) # печатаем переменную
                i = Image.open(filename)
                per = filename[14:18]
                print("печатаем значение срезанной - имя файла",per)

#блок считывания exif информации__________________________________________________________
               # i = Image.open(filename)
                info = i._getexif()
                print(info)
                pnom0 = str(info.get(315)[1])
                pnom = pnom0.replace('/', ' ').replace('-', ' ').replace('"', ' ').replace("'", ' ').replace(":", ' ').replace(
                        "&", ' ').replace("*", ' ').replace("b", ' ').replace(" ", '').replace("?", '').replace("\\", '')
                print('Печатаем порядковый номер фотографии:',pnom)
                i.close()
                for treshold_number in range(1, 254, 1):
                    img_grey = cv2.imread(filename, cv2.IMREAD_GRAYSCALE)
                    thresh = treshold_number
                    img_binary = cv2.threshold(img_grey, thresh, 255, cv2.THRESH_BINARY)[1]
                    print(f"печатаем значение тресхолд", thresh)
                    decoded = decode((img_binary), symbols=[ZBarSymbol.EAN13, ZBarSymbol.CODE128, ZBarSymbol.UPCA])
#вывод перебора фото на экран
                    def show():
                        cv2.namedWindow('custom window', cv2.WINDOW_KEEPRATIO)
                        cv2.imshow('custom window', img_binary)
                        cv2.resizeWindow('custom window', 1000, 1000)

                        cv2.waitKey(888)
                        cv2.destroyAllWindows()

                    #show()
                    if decoded:
                        break


                if decoded and pnom == o1: # если переменная decoded обнаружилась
                                print("печатаем цикл if decoded and pnom == o1")
                                print('Печатаем распознанный штрихкод:',decoded[0].data.decode('utf-8')) # печатаем переменную decoded в формате utf
                                x = str(decoded[0][0:19][0])
                                #z1 = str(x)
                                sheka = x.replace('/', ' ').replace('-', ' ').replace('"', ' ').replace("'", ' ').replace(":", ' ').replace(
                                    "&", ' ').replace("*", ' ').replace("b", ' ').replace(" ", '').replace("?", '')
                                #sheka = z2
                                guid_jpg = pnom + "_" + str(per) + '.jpg'

#блок поиска в таблице GUID номера_______________________________________________________________________________________
                                file = 'matrica/matrica_sent_stm.xlsx'
                                path_to_file = file
                                search_text = sheka
                                search_text = search_text.lower()
                                #print('Ищем в таблице строку содержащую:', search_text)

                                wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
                                sheets_list = wb.sheetnames  # Получаем список всех листов в файле
                                sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
                                row_max = sheet_active.max_row  # Получаем количество столбцов
                                # print(type(row_max))
                                column_max = sheet_active.max_column  # Получаем количество строк
                                #print('В файле:', path_to_file, '\n Cтолбцов:', row_max, '\n Колонок:', column_max)

                                row_min = 1  # Переменная, отвечающая за номер строки
                                column_min = 1  # Переменная, отвечающая за номер столбца
                                guid = str(random.randint(0, 10000))

                                while column_min <= column_max:
                                    row_min_min = row_min
                                    row_max_max = row_max
                                    while row_min_min <= row_max_max:
                                        row_min_min = str(row_min_min)

                                        word_column = get_column_letter(column_min)
                                        word_column = str(word_column)
                                        word_cell = word_column + row_min_min
                                        data_from_cell = sheet_active[word_cell].value

                                        data_from_cell = str(data_from_cell)
                                        regular = search_text
                                        result = re.findall(regular, data_from_cell)
                                        if len(result) > 0:
                                            #print('Нашли в ячейке:', word_cell)
                                            #print(word_cell)
                                            stroka = word_cell[:1]
                                            #print('строка товара:', stroka)
                                            stolbec = word_cell[1:]
                                            #print('столбец товара:', stolbec)
                                            stolbecnew = "A"

                                            guidyacheyka = stolbecnew + stolbec
                                            #print(guidyacheyka)
                                            guid: object = sheet_active[guidyacheyka].value
                                            print('Печатаем найденный гуид:',guid)
                                            guid_jpg = pnom + "_" + str(per) + '.jpg'
                                            print('Печатаем Guid_gpg:', guid_jpg)

                                        row_min_min = int(row_min_min)
                                        row_min_min = row_min_min + 1
                                    column_min = column_min + 1

#блок создания пути

                                fullPath = os.path.join(path, guid) # переменная fullpath равна
                                print("Печатаем fullPath", fullPath)
                                if not os.path.exists(fullPath): #возвращает True, если path указывает на существующий путь или дескриптор открытого файла
                                    os.mkdir(fullPath)# создает дирректорию fullpaht

#переносим файл в данную дирректорию
                                    print("создана директория",fullPath)
                                    if fullPath:
                                            print("none")
                                            a = filename
                                            b = fullPath
                                            print('Печатаем Путь куда 1:',fullPath)
                                            # shutil.copy(a, b)
                                            dest_dir = fullPath
                                            new_name = guid_jpg
                                            current_file_name = a
                                            os.rename(current_file_name, dest_dir + "/" + new_name)
                                            zoo = 1
                                            if zoo == 1:
                                                for filename in glob.glob("vhod00001/*." + "jpeg" or "jpg"):  # обходим заданную папку
                                                    if filename:  # если имя
                                                        str(filename)  # переводим имя в буквы
                                                        print('Печатаем распознанный путь после найденного ШК:', filename)  # печатаем переменную
                                                        per = filename[14:18]
                                                        print("печатаем значение срезанной - имя файла", per)

# блок считывания exif информации__________________________________________________________
                                                        i = Image.open(filename)
                                                        info = i._getexif()
                                                        print(info)
                                                        i.close()
                                                        pnom0 = info.get(315)[1]
                                                        pnom2 = str(pnom0)
                                                        pnom1 = pnom2.replace('/', ' ').replace('-', ' ').replace('"',
                                                                                                                  ' ').replace(
                                                            "'", ' ').replace(":", ' ').replace(
                                                            "&", ' ').replace("*", ' ').replace("b", ' ').replace(" ",
                                                                                                                  '').replace(
                                                            "?", '').replace("\\", '')
                                                        pnom = str(pnom1)
                                                        print('Печатаем порядковый номер фотографии:', pnom)

                                                        def pn_ne_odin():
                                                            str(filename)  # переводим имя в буквы
                                                            print("печатаем имя файла без шк",
                                                                  str(filename))  # печатаем переменную
                                                            a = filename
                                                            guid2 = guid + "_" + pnom + "_" + str(per) + '.jpg'
                                                            dest_dir = fullPath
                                                            #     per2 = per + 1
                                                            new_name = guid2
                                                            current_file_name = a
                                                            os.rename(current_file_name, dest_dir + "/" + new_name)
                                                            print('Печатаем порядковый номер фотографии:', pnom)

                                                        if pnom == o2:
                                                            pn_ne_odin()

                                                        if pnom == o3:
                                                            pn_ne_odin()

                                                        if pnom == o4:
                                                            pn_ne_odin()

                                                        if pnom == o5:
                                                            pn_ne_odin()

                                                        if pnom == o6:
                                                            pn_ne_odin()

                                                        if pnom == o7:
                                                            pn_ne_odin()

                                                        if pnom == o8:
                                                            pn_ne_odin()

                                                        if pnom == o9:
                                                            pn_ne_odin()

                                                        if pnom == o1:
                                                            o()

                elif not decoded and pnom == o1:
                            print("запускаеем цикл elif not decoded and pnom == o1")
                            filename2 = filename
                            if filename2:  # если найдено имя
                                str(filename2)  # переводим имя в строку
                                print("печатаем имя файла без шк", str(filename2))  # печатаем переменную
                                #a = filename2
                                cheredovanie2 =+1
                                fullPath = os.path.join(path, str(cheredovanie2))  # переменная fullpath равна
                                print("Печатаем fullPath", fullPath)
                                if not os.path.exists(
                                        fullPath):  # возвращает True, если path указывает на существующий путь или дескриптор открытого файла
                                    os.mkdir(fullPath)  # создает дирректорию fullpaht

                                guid2 = pnom + "_" + str(per) + '.jpg'
                                dest_dir = fullPath
                                new_name = guid2
                                current_file_name = filename2
                                os.rename(current_file_name, dest_dir + "/" + new_name)

                elif not decoded and pnom == o2:
                    print("elif not decoded and pnom == o2")
                    def not_decoded_and_bez_pnom():
                        filename2 = filename
                        if filename2:  # если найдено имя
                            str(filename2)  # переводим имя в буквы
                            #print("печатаем имя файла без шк", str(filename2))  # печатаем переменную
                            #a = filename2
                            guid2 = pnom + "_" + str(per) + '.jpg'
                            dest_dir = fullPath
                            new_name = guid2
                            current_file_name = filename2
                            os.rename(current_file_name, dest_dir + "/" + new_name)


                    not_decoded_and_bez_pnom()

                elif decoded and pnom == o2:
                    not_decoded_and_bez_pnom()

o()

#оповещение о завершении программы
endpic = Image.open("pic/endpic.png") #читаем и открываем изображение
endpic.show()#show image
playsound("sound/3.wav")#проигрываем музыку
